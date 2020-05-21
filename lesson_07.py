#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2020/5/20 13:35
# @Author : Lemon_Tricy
# @QQ: 2378807189
# Copyright：湖南省零檬信息技术有限公司

'''
1、用例，读取测试数据  ---- Done == read_data()
2、用数据发送接口请求，执行结果  --- Done  === post_func（）
3、执行结果 vs 预期结果  ==得出结论 ？？
4、得到的结论 会写到测试用例   ==Done  === write_result()
'''
import openpyxl
import requests
session = requests.session()   # requests库的session模块  复制给一个变量
# 读取测试用例数据的函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  # 加载了这个工作簿  -- Excel表格 ==赋值给一个变量
    sheet = wb[sheetname]   # 表单
    max_row = sheet.max_row  # 获取最大行号
    cases = []  # 空列表
    for i in range(2,max_row+1):
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,   # 获取url
        data = sheet.cell(row=i,column=6).value,  # 获取data
        expected_result = sheet.cell(row=i,column=7).value  # 获取期望结果
        )   # 一个用例存放到一个字典
        cases.append(case)   # 把字典追加到列表里保存起来
    return cases    # 定义成返回值
# 写入结果的方法--函数
def write_result(filename,sheetname,row,column,real_result):
    wb = openpyxl.load_workbook(filename)  # 加载了这个工作簿  -- Excel表格 ==赋值给一个变量
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = real_result   # 写入
    wb.save(filename)
# 函数-- 发送接口请求
def post_func(qcd_url,qcd_data):
    res = session.post(url=qcd_url,data=qcd_data)  # post方法发送接口请求  ==session来发送接口请求 ，自动带cookies值
    result = res.json()   # 字典
    return result   # 返回值 --- 响应消息结果

def execute_func(filename,sheetname):
    test_cases = read_data(filename,sheetname)   # 调用读取数据的函数
    for test_case in test_cases:
        case_id = test_case.get('case_id')   # 获取到对应case_id
        url = test_case.get('url')   # 获取到对应url
        data = test_case.get('data')   # 获取到对应参数   == 文本--字符串 ==转化为字典格式
        data = eval(data)  # eval()进行数据类型的转化  -- 字符串--> 字典
        expected_result = test_case['expected_result']   # 获取到对应期望结果  ==字符串
        expected_result = expected_result.replace('null','None')  # 字符串的替换
        expected_result = eval(expected_result)  #  eval()进行数据类型的转化  -- 字符串--> 字典
        real_result = post_func(qcd_url=url,qcd_data=data)   #调用接口发送函数 --接口请求
        # data --必须是字典格式
        real_msg = real_result.get('msg')  # 字典取值--获取要断言的有效字段
        expected_msg = expected_result.get('msg')
        print('真实执行结果是：{}'.format(real_msg))
        print('期望测试结果是：{}'.format(expected_msg))
        if real_msg == expected_msg:
            print('第{}条测试用例测试通过！'.format(case_id))
            final_result = 'Passed'   # 变量--目的： 会写结果
        else:
            print('第{}条测试用例测试不通过！'.format(case_id))
            final_result = 'Failed'
        print('**' * 20)
        write_result(filename,sheetname,case_id+1,8,final_result)    # 调用了会写的函数--结果写入

execute_func('test_case.xlsx','recharge')
# 充值需要传入登录的cookies --传入充值接口


